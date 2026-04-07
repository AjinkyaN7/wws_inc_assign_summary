import pandas as pd
import numpy as np
import requests, json
from datetime import date, timedelta
from datetime import datetime as dt
from openpyxl import load_workbook, Workbook, drawing, utils
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill, colors
import platform
from logzero import logger, logfile
import messaging as mg
import time
import os
import calendar
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.triggers.cron import CronTrigger

DEBUG = int(os.environ.get("DEBUG")) if os.environ.get("DEBUG") != None else None
if DEBUG == None:
    DEBUG = 0
    logger.warning("No DEBUG passed. Defalting to production mode.")

version = platform.python_version().split(".")[0]
if version == "3":
    import app_config.app_config as cfg
elif version == "2":
    import app_config as cfg
config = cfg.getconfig()

email = mg.Email()

CONFIG_URL = config["api"]["meta"]

filePath = "/src/uploads/mail/report-mail/wws/"
# filePath = "reportmail/report-mail/"
if "log" in os.listdir(os.getcwd()):
    logger.info("log directory exists")
else:
    logger.info("Creating log directory...")
    try:
        os.mkdir("log")
        logger.info("log directory created")
    except Exception as e:
        logger.exception("Unable to create log directory: " + str(e))
        exit()

if not (DEBUG):
    logfile(os.getcwd() + "/log/incSum.log", maxBytes=1e6, backupCount=5)
else:
    logfile(os.getcwd() + "/log/incSumDbg.log", maxBytes=1e6, backupCount=5)


# ---------------------------------------------------------------------------
# IST-based start time (st) and end time (et) calculation
# All times are returned as IST strings with 'T' and '.000Z' suffix to match
# the existing API format used by fetchIncidents.
# ---------------------------------------------------------------------------

def get_time_range(report_type):
    """
    Returns (st, et) as IST datetime strings for the API.

    Daily   : yesterday 00:00:00 IST  →  yesterday 23:59:59 IST
    Weekly  : previous Monday 00:00:00 IST →  previous Sunday 23:59:59 IST
    Monthly : previous month 1st 00:00:00 IST → previous month last day 23:59:59 IST
    """
    # Current time in IST (UTC + 5:30)
    ist_now = dt.utcnow() + timedelta(hours=5, minutes=30)
    today_ist = ist_now.date()

    if report_type == "daily":
        # Yesterday 12:00 AM to yesterday 11:59 PM
        yesterday = today_ist - timedelta(days=1)
        start_date = yesterday
        end_date = yesterday

    elif report_type == "weekly":
        # Previous week: Monday 12:00 AM to Sunday 11:59 PM
        # today_ist.weekday() == 0 means Monday (report runs on Monday)
        # Previous Sunday = today - 1, Previous Monday = today - 7
        prev_sunday = today_ist - timedelta(days=1)
        prev_monday = prev_sunday - timedelta(days=6)
        start_date = prev_monday
        end_date = prev_sunday

    elif report_type == "monthly":
        # Previous month: 1st 12:00 AM to last day 11:59 PM
        # (report runs on 1st of current month)
        last_day_prev_month = today_ist.replace(day=1) - timedelta(days=1)
        start_date = last_day_prev_month.replace(day=1)
        end_date = last_day_prev_month

    else:
        raise ValueError(f"Unknown report_type: {report_type}")

    st = str(start_date) + "T00:00:00.000Z"
    et = str(end_date) + "T23:59:59.000Z"

    logger.info(f"[{report_type.upper()}] Start time (IST): {st}    End time (IST): {et}")
    return st, et


def getUnitDet():
    url = (
        CONFIG_URL
        + '/customers?filter={"where":{"bu":"wws"},"fields":["id","name"],"include":{"relation":"sites","scope":{"fields":["id","name"],"include":{"relation":"units","scope":{"fields":["name","id"]}}}}}'
    )
    res = requests.get(url)
    if res.status_code == 200:
        custDet = json.loads(res.content)
        units = []
        unitsInfo = {}
        for cust in custDet:
            for site in cust["sites"]:
                for unit in site["units"]:
                    if "Test" in unit["name"] or "Demo" in unit["name"]:
                        pass
                    else:
                        units.append(unit["id"])
                        unitsInfo[unit["id"]] = {
                            "cust": cust["name"],
                            "site": site["name"],
                            "unit": unit["name"],
                        }
    else:
        logger.error("Units details retrieval failed  - " + str(res.status_code))
    logger.info("Number of units after deleting demo/test:" + str(len(units)))
    return units, unitsInfo


def fetchIncidents(units, st, et):
    fields = [
        "unitsId",
        "startTime",
        "comments",
        "assignedTo",
        "incidentName",
        "approved",
        "priority",
        "endTime",
        "id",
    ]
    w = {
        "unitsId": {"inq": units},
        "or": [{"open": True}, {"endTime": {"between": [st, et]}}],
    }
    url = CONFIG_URL + "/incidents/alter"
    postBody = {"fields": fields, "where": w}

    res1 = requests.post(url, json=postBody)
    if res1.status_code == 200:
        try:
            incs = json.loads(res1.content)["result"]
        except Exception as e:
            logger.exception("result not available in response - " + str(e))
            incs = []
    else:
        logger.error("Fetching incidents status code - " + str(res1.status_code))
        print(res1)
        incs = []

    return incs


def getReportDf(incs):
    data = pd.DataFrame(
        columns=[
            "Sno.",
            "Customer",
            "Site",
            "Unit",
            "Incident",
            "Criticality",
            "Start Time",
            "End Time",
            "Assignee",
            "Approved",
            "Expert Comment",
            "Expert RCA",
        ]
    )

    index = 0
    for inc in incs:
        assignees = []
        comm = []
        rca = []

        data.loc[index, "Customer"] = unitsInfo[inc["unitsId"]]["cust"]
        data.loc[index, "Site"] = unitsInfo[inc["unitsId"]]["site"]
        data.loc[index, "Unit"] = unitsInfo[inc["unitsId"]]["unit"]
        url = (
            config["api"]["public_datacenter_url"]
            + "#/incidents?incidentId="
            + inc["id"]
            + "&unitId="
            + inc["unitsId"]
        )
        data.loc[index, "Incident"] = '=HYPERLINK("{}", "{}")'.format(
            url, inc["incidentName"]
        )
        data.loc[index, "Start Time"] = (
            inc["startTime"][:10] + " " + inc["startTime"][11:-8]
        )

        try:
            data.loc[index, "End Time"] = (
                inc["endTime"][:10] + " " + inc["endTime"][11:-8]
            )
        except:
            data.loc[index, "End Time"] = "-"

        if "comments" in inc:
            for comment in inc["comments"]:
                if "type" in comment and comment["type"] == "expertComment":
                    comm.append(
                        comment["by"]
                        + " ("
                        + comment["time"][:10]
                        + " "
                        + comment["time"][11:-8]
                        + "): "
                        + (comment["comment"] if "comment" in comment else "-")
                    )
                if "type" in comment and comment["type"] == "expertRCA":
                    rca.append(
                        comment["by"]
                        + " ("
                        + comment["time"][:10]
                        + " "
                        + comment["time"][11:-8]
                        + "): "
                        + (comment["comment"] if "comment" in comment else "-")
                    )
            if len(comm) == 0:
                comm = ["-"]
            if len(rca) == 0:
                rca = ["-"]
        else:
            comm = ["-"]
            rca = ["-"]
        data.loc[index, "Expert Comment"] = ", ".join(comm)
        data.loc[index, "Expert RCA"] = ", ".join(rca)

        try:
            if len(inc["assignedTo"]) == 0:
                data.loc[index, "Assignee"] = "-"
            else:
                for assgn in inc["assignedTo"]:
                    assignees.append(
                        assgn["firstName"].capitalize()
                        + " "
                        + assgn["lastName"].capitalize()
                    )
                data.loc[index, "Assignee"] = ", ".join(list(set(assignees)))
        except:
            data.loc[index, "Assignee"] = "-"

        if inc["approved"]:
            data.loc[index, "Approved"] = "Yes"
        else:
            data.loc[index, "Approved"] = "No"

        if inc["priority"] == "danger":
            data.loc[index, "Criticality"] = "Critical"
        elif inc["priority"] == "warning":
            data.loc[index, "Criticality"] = "Warning"
        else:
            data.loc[index, "Criticality"] = "Deviation"

        index += 1

    return data


def prepAttachment(df, report_type):
    try:
        writer = pd.ExcelWriter(
            filePath
            + "IncidentAssgnSummary_"
            + report_type.capitalize()
            + "_"
            + (date.today() - timedelta(days=1)).strftime("%d%b%Y")
            + ".xlsx",
            engine="openpyxl",
        )
    except Exception as e:
        logger.exception("EXCEPTION: " + str(e))

    df.to_excel(writer, sheet_name="Sheet1", index=False)

    worksheet = writer.sheets["Sheet1"]

    bd = Side(style="thin", color="000000")
    for cell in list(worksheet._cells.values()):
        cell.font = Font(name="Verdana", size=10)
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(
        min_row=1, max_row=1, min_col=1, max_col=df.shape[1]
    ):
        for cell in row:
            cell.font = Font(name="Verdana", size=10, bold=True)
            cell.fill = PatternFill(
                fill_type="solid", start_color="00FF9900", end_color="00FF9900"
            )

    index = list(df.columns).index("Incident") + 1
    for i in range(2, len(df) + 2):
        worksheet.cell(row=i, column=index).hyperlink = worksheet.cell(
            row=i, column=index
        ).value.split(",")[0][12:-1]
        worksheet.cell(row=i, column=index).value = worksheet.cell(
            row=i, column=index
        ).value.split(",")[1][2:-2]
        worksheet.cell(row=i, column=index).style = "Hyperlink"
        worksheet.cell(row=i, column=index).font = Font(
            name="Verdana", size=10, u="single", color=colors.BLUE
        )
        worksheet.cell(row=i, column=index).border = Border(
            left=bd, top=bd, right=bd, bottom=bd
        )
        worksheet.cell(row=i, column=index).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                try:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
                except:
                    dims[cell.column_letter] = max(
                        (
                            dims.get(cell.column_letter, 0),
                            len(cell.value.encode("utf-8")),
                        )
                    )

    for col, value in list(dims.items()):
        if value > 90:
            colWidth = 90
        else:
            colWidth = value + 3
        worksheet.column_dimensions[col].width = colWidth

    worksheet.freeze_panes = "H2"
    writer.close()


def prepMailTxt(df, report_type):
    assgnSummary = df.Assignee.value_counts()

    # Label changes based on report type
    report_label_map = {
        "daily": "daily",
        "weekly": "weekly",
        "monthly": "monthly",
    }
    report_label = report_label_map.get(report_type, "daily")

    headers = ""
    for colName in ["Sno.", "Assignee", "Incidents Count", "Critical Incidents Count"]:
        headers += '<th style="background-color:#FFA000">' + colName + "</th>"

    contents = ""
    for i in range(len(assgnSummary)):
        if assgnSummary.index[i] == "-":
            txt = "Unassigned"
        else:
            txt = assgnSummary.index[i]
        contents += (
            '<tr><td style="text-align:center">'
            + str(i + 1)
            + "</td>"
            + '<td style="text-align:justify">'
            + txt
            + "</td>"
            + '<td style="text-align:center">'
            + str(assgnSummary[i])
            + "</td>"
            + '<td style="text-align:center">'
            + str(
                len(
                    df.loc[df["Assignee"] == assgnSummary.index[i]].loc[
                        df["Criticality"] == "Critical"
                    ]
                )
            )
            + "</td></tr>"
        )

    mailText = (
        "<html><head></head><body>Hi,<br><br>Please find below the "
        + report_label
        + " incident assignment summary. "
        + "The details are also attached with this mail.<br><br>"
        + '<table border= "1px solid black"><caption style="text-align:left"><b><u>Summary</u></b></caption>'
        + "<tr>"
        + headers
        + "</tr>"
        + contents
        + "</table><br><br>"
        + "<i>Please NOTE: This is a system generated e-mail, DO NOT REPLY</i></body></html>"
    )
    return mailText


def get_attachment_filename(report_type):
    return (
        filePath
        + "IncidentAssgnSummary_"
        + report_type.capitalize()
        + "_"
        + (date.today() - timedelta(days=1)).strftime("%d%b%Y")
        + ".xlsx"
    )


def mailer(text, sub, to, cc, st, report_type="daily"):
    logger.info("Within mailer for " + st)
    body = {
        "to": to,
        "subject": sub,
        "html": text,
        "cc": cc,
        "bcc": ["ajinkya.n@exactspace.co", "jyotiranjan.n@exactspace.co"],
    }
    if st == "Internal":
        try:
            res = email.sendSESMail(body)
            logger.info("Mail response: " + str(res))
        except Exception as e:
            res = "Mail error.."
            logger.exception("EXCEPTION: mail sending issue - " + str(e))
    else:
        body["f1"] = get_attachment_filename(report_type)
        try:
            res = email.sendSESMailWithAttach(body)
            logger.info("Mail response: " + str(res))
        except Exception as e:
            res = "Mail error.."
            logger.exception("EXCEPTION: mail sending issue - " + str(e))
    return res


units, unitsInfo = getUnitDet()


def prepReport(report_type="daily"):
    """
    Main report function. Accepts report_type: 'daily', 'weekly', 'monthly'.
    Computes IST-based st/et, fetches incidents, builds report, and emails.
    """
    rep = 0

    def _run():
        nonlocal rep
        if rep < 5:
            st, et = get_time_range(report_type)

            try:
                incs = fetchIncidents(units, st, et)
            except Exception as e:
                logger.exception("EXCEPTION: Fetching incidents - " + str(e))
                incs = []

            try:
                data = getReportDf(incs)
            except Exception as e:
                logger.exception("EXCEPTION: Converting inc to dataframe - " + str(e))
                data = pd.DataFrame()

            try:
                data = data.sort_values(
                    ["Assignee", "Criticality", "Customer", "Site", "Unit", "Start Time"],
                    ascending=[True, True, True, True, True, True],
                    ignore_index=True,
                )
                assgn = data.pop("Assignee")
                data.insert(1, "Assignee", assgn)
            except Exception as e:
                logger.exception("EXCEPTION: Initial dataframe prep - " + str(e))

            try:
                data1 = data[data["Assignee"] == "-"]
                data2 = data[data["Assignee"] != "-"]
                df = pd.concat([data2, data1])
                df["Sno."] = np.arange(1, len(df) + 1, 1)
                df = df.reset_index(drop=True)
            except Exception as e:
                logger.exception("EXCEPTION: Sorted dataframe prep - " + str(e))
                df = pd.DataFrame()

            if len(df) > 0:
                try:
                    prepAttachment(df, report_type)
                except Exception as e:
                    logger.exception("EXCEPTION: Attachment prep - " + str(e))

                try:
                    mailText = prepMailTxt(df, report_type)
                except Exception as e:
                    logger.exception("EXCEPTION: Mailer text prep - " + str(e))

                subject = (
                    "Incident Assignment Summary ("
                    + report_type.capitalize()
                    + ") - "
                    + (date.today() - timedelta(days=1)).strftime("%d%b%Y")
                )

                if DEBUG:
                    to = ["ajinkya.n@exactspace.co", "jyotiranjan.n@exactspace.co"]
                    cc = ["sreereddy.m@exactspace.co"]
                else:
                    to = ["ajinkya.n@exactspace.co","jyotiranjan.n@exactspace.co"]
                    cc = ["sreereddy.m@exactspace.co"]

                resp = mailer(mailText, subject, to, cc, "External", report_type)
                if "Success" in str(resp):
                    logger.info(f"[{report_type.upper()}] Mail sent successfully")
                    try:
                        os.remove(get_attachment_filename(report_type))
                    except Exception as e:
                        logger.exception(
                            "EXCEPTION: Unable to delete file after sending mail - "
                            + str(e)
                        )
                else:
                    rep += 1
                    time.sleep(5 * 60)
                    _run()
            else:
                rep += 1
                time.sleep(5 * 60)
                _run()
        else:
            mailText = "<html><head></head><body>Something problematic with incident assignment summary population...</body></html>"
            subject = f"Incident Assignment Summary ({report_type.capitalize()}) failed!!!"
            to = ["ajinkya.n@exactspace.co", "jyotiranjan.n@exactspace.co"]
            resp = mailer(mailText, subject, to, [], "Internal", report_type)

    _run()


# ---------------------------------------------------------------------------
# SCHEDULER SETUP
# Server runs in UTC. IST = UTC + 5:30
#
# Daily   : 10:00 AM IST  =  04:30 UTC  → every day
# Weekly  : 10:15 AM IST  =  04:45 UTC  → every Monday
# Monthly : 10:30 AM IST  =  05:00 UTC  → 1st of every month
# ---------------------------------------------------------------------------

scheduler = BlockingScheduler()

# Daily report — every day at 04:30 UTC (10:00 AM IST)
scheduler.add_job(
    prepReport,
    CronTrigger(hour=4, minute=30),
    kwargs={"report_type": "daily"},
    id="daily_report",
    name="Daily Incident Report",
)

# Weekly report — every Monday at 04:45 UTC (10:15 AM IST)
scheduler.add_job(
    prepReport,
    CronTrigger(day_of_week="mon", hour=4, minute=45),
    kwargs={"report_type": "weekly"},
    id="weekly_report",
    name="Weekly Incident Report",
)

# Monthly report — 1st of every month at 05:00 UTC (10:30 AM IST)
scheduler.add_job(
    prepReport,
    CronTrigger(day=1, hour=5, minute=0),
    kwargs={"report_type": "monthly"},
    id="monthly_report",
    name="Monthly Incident Report",
)

logger.info("Scheduler started. Waiting for jobs...")
logger.info("  Daily  : 04:30 UTC (10:00 AM IST) every day")
logger.info("  Weekly : 04:45 UTC (10:15 AM IST) every Monday")
logger.info("  Monthly: 05:00 UTC (10:30 AM IST) 1st of every month")

scheduler.start()